"""
Local Documents Crawler — France / Italy / Spain
=================================================
Reads university databases from Excel files in the Documents folder and
imports English-taught bachelor and master programs into masters_programs.

Sources:
  Documents/France University Database.xlsx  — Masters (sheet "France"), specific
    programs (sheet "France Masters"), Bachelors (cols 12-20 of "France" sheet)
  Documents/Italy University Database.xlsx   — Masters by university (sheet "Italy")
  Documents/Spain University Database.xlsx   — Masters + Bachelors (sheet "Spain")

Run:
  python programs_pipeline.py --source local_docs
  python programs_pipeline.py --source local_docs --dry-run
"""

import hashlib
import logging
import re
import time
from pathlib import Path
from typing import Optional

import httpx
from bs4 import BeautifulSoup

from base_program import BaseProgramCrawler, RawProgram

logger = logging.getLogger("local_docs")

DOCS_DIR = Path(__file__).parent.parent.parent / "Documents"
EUR_TO_USD = 1.09

# ── field/category helpers ────────────────────────────────────────────────────

ABBR_TO_FIELDS = {
    "Engg": "Engineering", "Arch": "Architecture", "CS": "Computer Science",
    "AI": "Artificial Intelligence", "Data": "Data Science", "Eco": "Economics",
    "Business": "Business", "Management": "Management", "Science": "Natural Sciences",
    "LifeSci": "Life Sciences", "BioSci": "Biological Sciences", "Physics": "Physics",
    "Math": "Mathematics", "Law": "Law", "Humanities": "Humanities",
    "Medicine": "Medicine", "Pharmacy": "Pharmacy", "Health": "Public Health",
    "Agriculture": "Agriculture", "FoodSci": "Food Science", "Forestry": "Forestry",
    "Tourism": "Tourism", "Social": "Social Sciences", "SocialSci": "Social Sciences",
    "Psychology": "Psychology", "Education": "Education", "Finance": "Finance",
    "Automotive": "Automotive Engineering", "CognitiveSci": "Cognitive Science",
    "Urban": "Urban Planning", "EnvSci": "Environmental Science",
    "Maritime": "Maritime Studies", "Bioinformatics": "Bioinformatics",
    "Linguistics": "Linguistics", "Languages": "Languages",
    "Economics": "Economics", "Neuroscience": "Neuroscience",
}

FIELD_TO_CATEGORY = {
    "Computer Science": "cs_ai", "Artificial Intelligence": "cs_ai",
    "Data Science": "cs_ai", "Bioinformatics": "cs_ai",
    "Engineering": "engineering", "Architecture": "engineering",
    "Automotive Engineering": "engineering",
    "Economics": "business", "Business": "business", "Management": "business",
    "Finance": "business",
    "Natural Sciences": "science", "Life Sciences": "science",
    "Biological Sciences": "science", "Physics": "science", "Mathematics": "science",
    "Environmental Science": "science",
    "Medicine": "health", "Public Health": "health", "Pharmacy": "health",
    "Law": "social", "Social Sciences": "social", "Psychology": "social",
    "Humanities": "social", "Political Science": "social", "Education": "social",
    "Cognitive Science": "social",
    "Architecture": "arts", "Urban Planning": "arts",
    "Agriculture": "agriculture", "Food Science": "agriculture", "Forestry": "agriculture",
    "Languages": "languages", "Linguistics": "languages",
}


def fields_from_abbr(raw: str) -> list[str]:
    """'Engg/Arch/CS' → ['Engineering', 'Architecture', 'Computer Science']"""
    if not raw:
        return []
    out = []
    for part in re.split(r"[/,&+]", str(raw)):
        part = part.strip()
        mapped = ABBR_TO_FIELDS.get(part)
        if mapped:
            out.append(mapped)
        elif len(part) > 3 and part.replace(" ", "").isalpha():
            out.append(part.title())
    return list(dict.fromkeys(out))[:4]


def category_from_fields(fields: list[str]) -> str:
    for f in fields:
        cat = FIELD_TO_CATEGORY.get(f)
        if cat:
            return cat
    return "cs_ai"


def category_from_name(name: str) -> str:
    n = name.lower()
    if any(k in n for k in ["computer", "software", "data", "ai ", "artificial", "cyber", "informatics"]):
        return "cs_ai"
    if any(k in n for k in ["engineering", "electrical", "mechanical", "civil", "chemical", "aerospace", "telecom"]):
        return "engineering"
    if any(k in n for k in ["business", "management", "finance", "economics", "mba", "marketing", "commerce", "international business"]):
        return "business"
    if any(k in n for k in ["physics", "chemistry", "biology", "mathematics", "science", "bioinformatics"]):
        return "science"
    if any(k in n for k in ["medicine", "health", "nursing", "pharmacy", "dental"]):
        return "health"
    if any(k in n for k in ["law", "political", "sociology", "psychology", "international relation"]):
        return "social"
    if any(k in n for k in ["architecture", "design", "art", "urban", "hospitality", "tourism", "culinary", "hotel"]):
        return "arts"
    if any(k in n for k in ["agriculture", "food", "forestry", "rural"]):
        return "agriculture"
    if any(k in n for k in ["language", "linguist", "italian", "french", "spanish", "german"]):
        return "languages"
    return "cs_ai"


def parse_tuition(raw) -> Optional[float]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("none", "free", "0", "-", "n/a"):
        return None
    nums = re.findall(r"[\d]+", s.replace(",", ""))
    if not nums:
        return None
    try:
        amount = float(nums[0])
        if amount < 100:
            return None
        return amount  # store as-is in EUR (all programs are European)
    except ValueError:
        return None


def parse_ielts(req: str) -> Optional[float]:
    if not req:
        return None
    s = str(req)
    m = re.search(r"IELTS\s*([\d.]+)", s, re.I)
    if m:
        return float(m.group(1))
    if re.search(r"\bB1\b", s):
        return 5.5
    if re.search(r"\bB2\b", s):
        return 6.0
    if re.search(r"\bC1\b", s):
        return 7.0
    m2 = re.search(r"TOEFL\s*i?BT\s*(\d+)", s, re.I)
    if m2:
        score = int(m2.group(1))
        if score >= 100:
            return 7.0
        if score >= 90:
            return 6.5
        return 6.0
    return None


def parse_duration(raw) -> float:
    if raw is None:
        return 2.0
    s = str(raw).lower()
    m = re.search(r"(\d+)\s*[-–]\s*(\d+)\s*year", s)
    if m:
        return float(m.group(1))
    m = re.search(r"(\d+)\s*year", s)
    if m:
        return float(m.group(1))
    return 2.0


def parse_deadline(raw) -> Optional[str]:
    if raw is None:
        return None
    import datetime
    if isinstance(raw, datetime.datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    # Try "March", "Jan–May", etc. — return None (rolling)
    if re.search(r"[A-Za-z]{3,}", s):
        return None
    m = re.search(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    return None


def make_fp(name: str, university: str, country: str, level: str) -> str:
    key = f"{name.lower().strip()}|{university.lower().strip()}|{country.lower()}|{level}"
    return hashlib.sha256(key.encode()).hexdigest()


def _skip_row(val) -> bool:
    """True if this looks like a section header or empty row."""
    if val is None:
        return True
    s = str(val).strip()
    if not s:
        return True
    skip_patterns = [
        r"^(region|community|🟦|🟨|🟥|lombardia|piemonte|veneto|toscana|lazio|puglia|campania|sicilia|sardegna)",
        r"^(île-de-france|catalonia|andalusia|aragon|galicia|navarre|basque|castile|asturias|cantabria|canary|balearic|valencian|murcia|north|central|south)",
        r"masters in \w+ \|",
        r"bachelors in \w+ \|",
        r"language program",
        r"^university$",
    ]
    for pat in skip_patterns:
        if re.search(pat, s, re.I):
            return True
    return False


# ── web enrichment helpers ────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}


def scrape_polimi_programs(client: httpx.Client) -> list[dict]:
    """Fetch Politecnico di Milano English master programs."""
    programs = []
    try:
        resp = client.get(
            "https://www.polimi.it/en/education/laurea-magistrale-programmes",
            timeout=20
        )
        if resp.status_code != 200:
            return []
        soup = BeautifulSoup(resp.text, "html.parser")
        for item in soup.select("li.course-item, div.course-title, h3.course-name, a[href*='/en/education/']"):
            name = item.get_text(strip=True)
            href = item.get("href", "")
            if len(name) > 10 and ("master" in name.lower() or "science" in name.lower() or
                                    "engineering" in name.lower() or "management" in name.lower()):
                programs.append({"name": name, "url": href if href.startswith("http") else
                                  "https://www.polimi.it" + href})
        if not programs:
            # Try heading-based extraction
            for tag in soup.find_all(["h2", "h3", "h4", "li"]):
                text = tag.get_text(strip=True)
                if 15 < len(text) < 120 and any(w in text.lower() for w in
                    ["master", "science", "engineering", "management", "design", "architecture"]):
                    a = tag.find("a")
                    url = ("https://www.polimi.it" + a["href"]) if a and a.get("href") else ""
                    programs.append({"name": text, "url": url})
    except Exception as e:
        logger.debug(f"Polimi scrape failed: {e}")
    return programs[:40]


def scrape_unibo_programs(client: httpx.Client) -> list[dict]:
    """Fetch Università di Bologna English master programs."""
    programs = []
    try:
        resp = client.get(
            "https://www.unibo.it/en/study/second-cycle-degree?lingua=en&intconmul=true",
            timeout=20
        )
        if resp.status_code != 200:
            return []
        soup = BeautifulSoup(resp.text, "html.parser")
        for item in soup.select("h3 a, .course-list a, li.course a, .programme-title a"):
            name = item.get_text(strip=True)
            href = item.get("href", "")
            if len(name) > 8:
                url = href if href.startswith("http") else "https://www.unibo.it" + href
                programs.append({"name": name, "url": url})
        if not programs:
            for tag in soup.find_all("a", href=re.compile(r"/en/study/")):
                name = tag.get_text(strip=True)
                if 10 < len(name) < 100:
                    url = tag["href"] if tag["href"].startswith("http") else "https://www.unibo.it" + tag["href"]
                    programs.append({"name": name, "url": url})
    except Exception as e:
        logger.debug(f"UniBO scrape failed: {e}")
    return programs[:40]


def scrape_polito_programs(client: httpx.Client) -> list[dict]:
    """Fetch Politecnico di Torino English master programs."""
    programs = []
    try:
        resp = client.get(
            "https://www.polito.it/en/education/master-s-degree-programmes",
            timeout=20
        )
        if resp.status_code != 200:
            return []
        soup = BeautifulSoup(resp.text, "html.parser")
        for a in soup.find_all("a", href=re.compile(r"/en/education/|/en/course/")):
            name = a.get_text(strip=True)
            if 10 < len(name) < 100:
                href = a["href"]
                url = href if href.startswith("http") else "https://www.polito.it" + href
                programs.append({"name": name, "url": url})
    except Exception as e:
        logger.debug(f"Polito scrape failed: {e}")
    return programs[:40]


def scrape_sapienza_programs(client: httpx.Client) -> list[dict]:
    """Fetch Sapienza Roma English master programs."""
    programs = []
    try:
        resp = client.get(
            "https://corsidilaurea.uniroma1.it/en",
            timeout=20
        )
        if resp.status_code != 200:
            return []
        soup = BeautifulSoup(resp.text, "html.parser")
        for a in soup.find_all("a", href=re.compile(r"/en/")):
            name = a.get_text(strip=True)
            if 10 < len(name) < 120 and any(w in name.lower() for w in
                ["master", "science", "engineering", "management", "law", "economics", "medicine"]):
                href = a["href"]
                url = href if href.startswith("http") else "https://corsidilaurea.uniroma1.it" + href
                programs.append({"name": name, "url": url})
    except Exception as e:
        logger.debug(f"Sapienza scrape failed: {e}")
    return programs[:40]


WEB_ENRICHMENT = {
    "Politecnico di Milano": ("Italy", "Milan", scrape_polimi_programs),
    "Università di Bologna": ("Italy", "Bologna", scrape_unibo_programs),
    "Politecnico di Torino": ("Italy", "Turin", scrape_polito_programs),
    "Sapienza Università di Roma": ("Italy", "Rome", scrape_sapienza_programs),
}


# ── PDF hardcoded data (English_Masters_France_2025_26 + Campus France List) ──

PDF_UNIVERSITIES: dict[str, tuple] = {
    # key: (name, city, qs_rank, tuition_eur, ielts, apply_url)
    # URLs verified against official university websites + Campus France database
    "psl":            ("Université PSL", "Paris", 28, 254, 6.5,
                       "https://psl.eu/en/education/find-your-curriculum/masters-degrees"),
    "ip_paris":       ("Institut Polytechnique de Paris", "Palaiseau", 41, 628, 6.5,
                       "https://www.ip-paris.fr/en/education/graduate-programs/masters-science"),
    "paris_saclay":   ("Université Paris-Saclay", "Orsay", 71, 254, 6.5,
                       "https://www.universite-paris-saclay.fr/en/education/masters-programmes-taught-english"),
    "sorbonne":       ("Sorbonne University", "Paris", 72, 254, 6.5,
                       "https://www.sorbonne-universite.fr/en/programs-english"),
    "sciences_po":    ("Sciences Po Paris", "Paris", 342, 14800, 7.0,
                       "https://www.sciencespo.fr/en/prospective-students/programmes-in-english/"),
    "paris_cite":     ("Université Paris Cité", "Paris", 300, 254, 6.5,
                       "https://u-paris.fr/language/en/programs-taught-in-english/"),
    "paris1":         ("Université Paris 1 Panthéon-Sorbonne", "Paris", 257, 254, 6.5,
                       "https://international.pantheonsorbonne.fr/en/join-paris-1-pantheon-sorbonne/studying-english"),
    "grenoble_alpes": ("Université Grenoble Alpes", "Grenoble", 321, 254, 6.5,
                       "https://www.univ-grenoble-alpes.fr/education/programs/programs-in-english-786070.kjsp"),
    "bordeaux":       ("Université de Bordeaux", "Bordeaux", 331, 254, 6.5,
                       "https://www.u-bordeaux.fr/en/international/come-to-bordeaux/international-students/international-study-programmes/masters-english"),
    "strasbourg":     ("Université de Strasbourg", "Strasbourg", 351, 254, 6.5,
                       "https://en.unistra.fr/studies/degree-programs"),
    "lyon1":          ("Université Claude Bernard Lyon 1", "Lyon", 481, 254, 6.5,
                       "https://www.univ-lyon1.fr/universite/international/degree-programs-taught-in-english"),
    # Grandes Écoles — business
    "insead":  ("INSEAD", "Fontainebleau", None, 89000, 7.0,
                "https://www.insead.edu/master-programmes"),
    "hec":     ("HEC Paris", "Jouy-en-Josas", None, 47000, 7.0,
                "https://www.hec.edu/en/master-s-programs"),
    "essec":   ("ESSEC Business School", "Cergy", None, 30000, 7.0,
                "https://www.essec.edu/en/programs/masters-msc/"),
    "escp":    ("ESCP Business School", "Paris", None, 18000, 7.0,
                "https://escp.eu/programmes/specialised-masters-MSc"),
    "edhec":   ("EDHEC Business School", "Lille", None, 20000, 7.0,
                "https://www.edhec.edu/en/programmes/masters-degree"),
    "emlyon":  ("emlyon Business School", "Lyon", None, 16000, 6.5,
                "https://em-lyon.com/en/masters-programs"),
    "neoma":   ("NEOMA Business School", "Reims", None, 14000, 6.5,
                "https://neoma-bs.com/masters-of-science"),
    "gem":     ("Grenoble École de Management", "Grenoble", None, 14000, 6.5,
                "https://www.grenoble-em.com/en/advanced-master-s-programs"),
}

# (univ_key, program_name, fields_list, category)
PDF_PROGRAMS: list[tuple] = [
    # ── PSL (QS #28) ─────────────────────────────────────────────────────────
    ("psl", "MSc Data Science", ["Data Science", "Statistics", "Machine Learning"], "cs_ai"),
    ("psl", "MSc Artificial Intelligence", ["Artificial Intelligence", "Machine Learning"], "cs_ai"),
    ("psl", "MSc Physics", ["Physics"], "science"),
    ("psl", "MSc Mathematics & Applications", ["Mathematics", "Applied Mathematics"], "science"),
    ("psl", "MSc Economics", ["Economics"], "business"),
    ("psl", "MSc Finance", ["Finance"], "business"),
    ("psl", "MSc Cognitive Science", ["Cognitive Science", "Neuroscience"], "social"),
    ("psl", "Master in Management", ["Management", "Business"], "business"),
    # ── IP Paris (QS #41) ────────────────────────────────────────────────────
    ("ip_paris", "MSc AI & Advanced Visual Computing", ["Artificial Intelligence", "Computer Vision"], "cs_ai"),
    ("ip_paris", "MSc Data & Decision Sciences", ["Data Science", "Statistics"], "cs_ai"),
    ("ip_paris", "MSc Cybersecurity", ["Cybersecurity", "Computer Science"], "cs_ai"),
    ("ip_paris", "MSc Machine Learning & Data Science", ["Machine Learning", "Data Science"], "cs_ai"),
    ("ip_paris", "MSc Economics & Public Policy", ["Economics", "Public Policy"], "business"),
    ("ip_paris", "MSc Advanced Wireless Communications", ["Telecommunications", "Electrical Engineering"], "engineering"),
    ("ip_paris", "MSc Sustainable Energy", ["Energy Engineering", "Sustainability"], "engineering"),
    # ── Paris-Saclay (QS #71, €254/yr) ───────────────────────────────────────
    ("paris_saclay", "MSc Artificial Intelligence", ["Artificial Intelligence", "Machine Learning"], "cs_ai"),
    ("paris_saclay", "MSc Data Knowledge & Hybrid AI", ["Data Science", "Artificial Intelligence"], "cs_ai"),
    ("paris_saclay", "MSc Distributed & Parallel Computing", ["Computer Science", "Systems Engineering"], "cs_ai"),
    ("paris_saclay", "MSc Human-Computer Interaction", ["Computer Science", "HCI"], "cs_ai"),
    ("paris_saclay", "MSc Economics & International Business", ["Economics", "International Business"], "business"),
    ("paris_saclay", "MSc Climate & Ecosystem Services", ["Environmental Science", "Climate Science"], "science"),
    ("paris_saclay", "MSc Advanced Wireless Communications", ["Telecommunications", "Signal Processing"], "engineering"),
    ("paris_saclay", "MSc Computational Neurosciences", ["Neuroscience", "Computational Biology"], "science"),
    ("paris_saclay", "MSc Bioinformatics", ["Bioinformatics", "Computational Biology"], "cs_ai"),
    # ── Sorbonne University (QS #72) ─────────────────────────────────────────
    ("sorbonne", "MSc Computer Science", ["Computer Science", "Software Engineering"], "cs_ai"),
    ("sorbonne", "MSc Bioinformatics", ["Bioinformatics", "Computational Biology"], "cs_ai"),
    ("sorbonne", "MSc Mathematics", ["Mathematics", "Applied Mathematics"], "science"),
    ("sorbonne", "MSc Physics", ["Physics"], "science"),
    ("sorbonne", "MSc Marine Science", ["Marine Science", "Environmental Science"], "science"),
    ("sorbonne", "MSc Economics", ["Economics"], "business"),
    # ── Sciences Po (QS #342) ─────────────────────────────────────────────────
    ("sciences_po", "MA International Affairs", ["International Relations", "Political Science"], "social"),
    ("sciences_po", "MA Public Policy", ["Public Policy", "Political Science"], "social"),
    ("sciences_po", "MA International Economic Policy", ["Economics", "International Relations"], "social"),
    ("sciences_po", "MA European Affairs", ["European Studies", "Political Science"], "social"),
    ("sciences_po", "Master in Finance & Strategy", ["Finance", "Strategy"], "business"),
    ("sciences_po", "MSc Data Science for Social Sciences", ["Data Science", "Social Science"], "cs_ai"),
    # ── Université Paris Cité (QS #300) ──────────────────────────────────────
    ("paris_cite", "MSc Biomedical Sciences", ["Biomedical Sciences", "Medicine"], "health"),
    ("paris_cite", "MSc Bioinformatics", ["Bioinformatics", "Computational Biology"], "cs_ai"),
    ("paris_cite", "MSc Computer Science", ["Computer Science"], "cs_ai"),
    ("paris_cite", "MSc Mathematics & Applications", ["Mathematics", "Applied Mathematics"], "science"),
    ("paris_cite", "MSc Urban Planning & Sustainable Development", ["Urban Planning", "Sustainability"], "arts"),
    # ── Université Paris 1 Panthéon-Sorbonne (QS #257) ───────────────────────
    ("paris1", "MSc International Business", ["International Business", "Economics"], "business"),
    ("paris1", "MSc Economics & Econometrics", ["Economics", "Econometrics"], "business"),
    ("paris1", "MSc Law", ["Law"], "social"),
    # ── Université Grenoble Alpes (QS #321) ──────────────────────────────────
    ("grenoble_alpes", "MSc Embedded Systems", ["Embedded Systems", "Electrical Engineering"], "engineering"),
    ("grenoble_alpes", "MSc Information Security", ["Cybersecurity", "Computer Science"], "cs_ai"),
    ("grenoble_alpes", "MSc Physics & Nanosciences", ["Physics", "Nanotechnology"], "science"),
    ("grenoble_alpes", "MSc Industrial & Applied Mathematics", ["Mathematics", "Industrial Engineering"], "science"),
    ("grenoble_alpes", "MSc Environmental Sciences", ["Environmental Science", "Earth Sciences"], "science"),
    # ── Université de Bordeaux (QS #331) ─────────────────────────────────────
    ("bordeaux", "MSc Computer Science", ["Computer Science"], "cs_ai"),
    ("bordeaux", "MSc Public Health", ["Public Health", "Epidemiology"], "health"),
    ("bordeaux", "MSc International Economics", ["Economics", "International Trade"], "business"),
    ("bordeaux", "MSc Neuroscience", ["Neuroscience", "Biology"], "science"),
    # ── Université de Strasbourg (QS #351) ───────────────────────────────────
    ("strasbourg", "MSc Photonics & Optical Engineering", ["Photonics", "Optical Engineering"], "engineering"),
    ("strasbourg", "MSc Bioinformatics", ["Bioinformatics", "Computational Biology"], "cs_ai"),
    ("strasbourg", "MSc European Law", ["Law", "European Law"], "social"),
    ("strasbourg", "MSc Chemistry", ["Chemistry", "Materials Science"], "science"),
    # ── Université Claude Bernard Lyon 1 (QS #481) ───────────────────────────
    ("lyon1", "MSc Biosciences", ["Biology", "Life Sciences"], "science"),
    ("lyon1", "MSc Mathematics", ["Mathematics"], "science"),
    ("lyon1", "MSc Computer Science", ["Computer Science"], "cs_ai"),
    # ── HEC Paris ─────────────────────────────────────────────────────────────
    ("hec", "MSc Marketing", ["Marketing", "Business"], "business"),
    ("hec", "MSc Finance", ["Finance"], "business"),
    ("hec", "MSc Strategy & International Business", ["Strategy", "International Business"], "business"),
    ("hec", "Master in Management (Grande École)", ["Management", "Business"], "business"),
    # ── INSEAD ────────────────────────────────────────────────────────────────
    ("insead", "MBA", ["Business Administration", "Management"], "business"),
    ("insead", "Master in Finance", ["Finance", "Investment"], "business"),
    # ── ESSEC Business School ─────────────────────────────────────────────────
    ("essec", "MSc Management", ["Management", "Business"], "business"),
    ("essec", "MSc Finance", ["Finance"], "business"),
    ("essec", "MSc Data Science & AI for Business", ["Data Science", "Business Analytics"], "cs_ai"),
    ("essec", "MSc Marketing & Business Development", ["Marketing", "Business"], "business"),
    # ── ESCP Business School ──────────────────────────────────────────────────
    ("escp", "Master in Management", ["Management", "Business"], "business"),
    ("escp", "MSc Marketing & Creativity", ["Marketing", "Creative Industries"], "business"),
    ("escp", "MSc Finance", ["Finance"], "business"),
    ("escp", "MSc Digital Transformation", ["Digital Business", "Technology Management"], "business"),
    # ── EDHEC Business School ─────────────────────────────────────────────────
    ("edhec", "MSc Finance", ["Finance"], "business"),
    ("edhec", "MSc Risk & Investment Management", ["Finance", "Risk Management"], "business"),
    ("edhec", "MSc Data Analytics & AI", ["Data Science", "Artificial Intelligence"], "cs_ai"),
    ("edhec", "MSc Corporate Finance & Banking", ["Finance", "Banking"], "business"),
    # ── emlyon Business School ────────────────────────────────────────────────
    ("emlyon", "MSc International Management", ["International Management", "Business"], "business"),
    ("emlyon", "MSc Data Science for Management", ["Data Science", "Business Analytics"], "cs_ai"),
    ("emlyon", "MSc Innovation, Strategy & Entrepreneurship", ["Innovation", "Entrepreneurship"], "business"),
    # ── NEOMA Business School ─────────────────────────────────────────────────
    ("neoma", "MSc International Finance", ["Finance", "International Business"], "business"),
    ("neoma", "MSc Management & Business Analytics", ["Business Analytics", "Management"], "cs_ai"),
    ("neoma", "MSc Digital Marketing", ["Digital Marketing", "Marketing"], "business"),
    # ── Grenoble École de Management ─────────────────────────────────────────
    ("gem", "MSc Management of Innovation & Industrial Systems", ["Innovation Management", "Engineering"], "business"),
    ("gem", "MSc Data Science & AI for Management", ["Data Science", "Business Analytics"], "cs_ai"),
    ("gem", "MSc International Business", ["International Business", "Management"], "business"),
]

# (name, city, tuition_eur, ielts, listing_url, [(prog_name, fields, category, specific_url), ...])
# listing_url = school's English programs catalog page (verified working)
# specific_url = direct page for that program (if known), else falls back to listing_url
ENGINEERING_SCHOOLS: list[tuple] = [
    ("CentraleSupélec", "Gif-sur-Yvette", 628, 6.5,
     "https://studyatcentralesupelec.fr/study-programmes-in-english/", [
         ("MSc Electrical Engineering", ["Electrical Engineering", "Electronics"], "engineering",
          "https://studyatcentralesupelec.fr/study-programmes-in-english/"),
         ("MSc Industrial Engineering & Systems", ["Industrial Engineering", "Systems Engineering"], "engineering",
          "https://studyatcentralesupelec.fr/study-programmes-in-english/"),
         ("MSc Applied Mathematics & Computer Science", ["Applied Mathematics", "Computer Science"], "cs_ai",
          "https://studyatcentralesupelec.fr/study-programmes-in-english/"),
     ]),
    ("École Centrale de Lyon", "Écully", 628, 6.5,
     "https://www.ec-lyon.fr/en/academics/master-degrees/international-master", [
         ("MSc Electrical Engineering", ["Electrical Engineering"], "engineering",
          "https://www.ec-lyon.fr/en/academics/master-degrees/international-master"),
         ("MSc Mechanical Engineering", ["Mechanical Engineering"], "engineering",
          "https://www.ec-lyon.fr/en/academics/master-degrees/international-master"),
     ]),
    ("École Centrale de Nantes", "Nantes", 628, 6.5,
     "https://www.ec-nantes.fr/english-version/study/our-english-taught-masters-programmes", [
         ("MSc Advanced Robotics", ["Robotics", "Mechanical Engineering"], "engineering",
          "https://www.ec-nantes.fr/english-version/study/our-english-taught-masters-programmes"),
         ("MSc Civil Engineering", ["Civil Engineering", "Construction"], "engineering",
          "https://www.ec-nantes.fr/english-version/study/our-english-taught-masters-programmes"),
     ]),
    ("École Centrale de Lille", "Lille", 628, 6.5,
     "https://centralelille.fr/en/formation/masters/", [
         ("MSc Industrial Engineering", ["Industrial Engineering"], "engineering",
          "https://centralelille.fr/en/formation/masters/"),
         ("MSc Computer Science & Data Science", ["Computer Science", "Data Science"], "cs_ai",
          "https://centralelille.fr/en/formation/masters/"),
     ]),
    ("Centrale Méditerranée", "Marseille", 628, 6.5,
     "https://www.centrale-mediterranee.fr/en/education/masters/international-masters", [
         ("MSc Mechanical Engineering", ["Mechanical Engineering"], "engineering",
          "https://www.centrale-mediterranee.fr/en/education/masters/international-masters"),
         ("MSc Biomedical Engineering", ["Biomedical Engineering", "Chemical Engineering"], "engineering",
          "https://www.centrale-mediterranee.fr/en/education/masters/master-biomedical-engineering"),
     ]),
    ("Arts et Métiers ParisTech", "Paris", 628, 6.5,
     "https://artsetmetiers.fr/en/programmes-taught-english", [
         ("MSc Mechanical Engineering", ["Mechanical Engineering", "Manufacturing"], "engineering",
          "https://artsetmetiers.fr/en/programmes-taught-english"),
         ("MSc Industrial Systems Engineering", ["Industrial Engineering", "Systems Engineering"], "engineering",
          "https://artsetmetiers.fr/en/formation/masters-programme-2-years-graduate"),
     ]),
    ("Télécom Paris", "Palaiseau", 628, 6.5,
     "https://www.telecom-paris.fr/en/masters/ip-paris", [
         ("MSc Cybersecurity", ["Cybersecurity", "Computer Science"], "cs_ai",
          "https://www.telecom-paris.fr/en/masters/ip-paris"),
         ("MSc Data & Artificial Intelligence", ["Data Science", "Artificial Intelligence"], "cs_ai",
          "https://www.telecom-paris.fr/en/masters/ip-paris"),
         ("MSc Networks & Telecommunications", ["Telecommunications", "Computer Networks"], "engineering",
          "https://www.telecom-paris.fr/en/masters/ip-paris"),
     ]),
    ("Télécom SudParis", "Évry", 628, 6.5,
     "https://www.telecom-sudparis.eu/en/formation/masters-of-science/", [
         ("MSc Cybersecurity", ["Cybersecurity", "Computer Science"], "cs_ai",
          "https://www.telecom-sudparis.eu/en/formation/masters-of-science/"),
         ("MSc Networks & Services", ["Telecommunications", "Computer Networks"], "engineering",
          "https://www.telecom-sudparis.eu/en/formation/masters-of-science/"),
     ]),
    ("INSA Lyon", "Lyon", 628, 6.5,
     "https://www.insa-lyon.fr/en/courses-taught-english", [
         ("MSc Computer Science & Software Engineering", ["Computer Science", "Software Engineering"], "cs_ai",
          "https://www.insa-lyon.fr/en/courses-taught-english"),
         ("MSc Biosciences Engineering", ["Biosciences", "Biomedical Engineering"], "science",
          "https://www.insa-lyon.fr/en/courses-taught-english"),
         ("MSc Civil Engineering & Urban Planning", ["Civil Engineering", "Urban Planning"], "engineering",
          "https://www.insa-lyon.fr/en/courses-taught-english"),
     ]),
    ("INSA Rennes", "Rennes", 628, 6.5,
     "https://www.insa-rennes.fr/en/courses/courses-taught-in-english.html", [
         ("MSc Electronics & Computer Engineering", ["Electronics", "Computer Engineering"], "engineering",
          "https://www.insa-rennes.fr/en/courses/courses-taught-in-english.html"),
         ("MSc Mechanical Engineering", ["Mechanical Engineering"], "engineering",
          "https://www.insa-rennes.fr/en/courses/courses-taught-in-english.html"),
     ]),
    ("INSA Strasbourg", "Strasbourg", 628, 6.5,
     "https://www.insa-strasbourg.fr/en/research-masters-degrees/", [
         ("MSc Civil Engineering", ["Civil Engineering", "Structural Engineering"], "engineering",
          "https://www.insa-strasbourg.fr/en/research-masters-degrees/"),
         ("MSc Mechanical Engineering", ["Mechanical Engineering"], "engineering",
          "https://www.insa-strasbourg.fr/en/research-masters-degrees/"),
     ]),
    ("INSA Toulouse", "Toulouse", 628, 6.5,
     "https://www.insa-toulouse.fr/en/masters-programs/", [
         ("MSc Aerospace Engineering", ["Aerospace Engineering", "Aeronautics"], "engineering",
          "https://www.insa-toulouse.fr/en/masters-programs/"),
         ("MSc Computer Science", ["Computer Science"], "cs_ai",
          "https://www.insa-toulouse.fr/en/masters-programs/"),
     ]),
    ("INSA Rouen Normandie", "Rouen", 628, 6.5,
     "https://www.insa-rouen.fr/en/education/masters", [
         ("MSc Chemical Engineering", ["Chemical Engineering", "Energy"], "engineering",
          "https://www.insa-rouen.fr/en/education/masters"),
         ("MSc Mechanical Engineering", ["Mechanical Engineering"], "engineering",
          "https://www.insa-rouen.fr/en/education/masters"),
     ]),
    ("ENAC", "Toulouse", 628, 6.5,
     "https://www.enac.fr/en/masters", [
         ("MSc Aviation & Transport", ["Aviation", "Transport Engineering"], "engineering",
          "https://www.enac.fr/en/masters"),
         ("MSc Aerospace Systems", ["Aerospace Engineering", "Systems Engineering"], "engineering",
          "https://www.enac.fr/en/admissions-concours/admissions-masters"),
     ]),
    ("ENSAE Paris", "Palaiseau", 628, 6.5,
     "https://www.ensae.fr/en/education/advanced-masters", [
         ("MSc Statistics & Econometrics", ["Statistics", "Econometrics"], "science",
          "https://www.ensae.fr/en/education/advanced-masters"),
         ("MSc Data Science", ["Data Science", "Machine Learning"], "cs_ai",
          "https://www.ensae.fr/en/education/advanced-masters"),
         ("MSc Finance & Insurance", ["Finance", "Actuarial Science"], "business",
          "https://www.ensae.fr/en/education/advanced-masters"),
     ]),
    ("ENSTA Paris", "Palaiseau", 628, 6.5,
     "https://www.ensta-paris.fr/en/training/other-programs/masters-programs", [
         ("MSc Sustainable Energy Systems", ["Energy Engineering", "Sustainability"], "engineering",
          "https://www.ensta-paris.fr/en/training/other-programs/masters-programs"),
         ("MSc Autonomous Vehicles & AI", ["Artificial Intelligence", "Robotics"], "cs_ai",
          "https://www.ensta-paris.fr/en/training/other-programs/masters-programs"),
         ("MSc Computer Science & Networks", ["Computer Science", "Networking"], "cs_ai",
          "https://www.ensta-paris.fr/en/training/other-programs/masters-programs"),
     ]),
    ("EURECOM", "Sophia Antipolis", 628, 6.5,
     "https://www.eurecom.fr/en/teaching", [
         ("MSc Security Science & Technology", ["Cybersecurity", "Computer Science"], "cs_ai",
          "https://www.eurecom.fr/en/teaching/master-computer-science"),
         ("MSc Data Science & Engineering", ["Data Science", "Computer Engineering"], "cs_ai",
          "https://www.eurecom.fr/en/teaching/master-computer-science"),
         ("MSc Communication Systems", ["Telecommunications", "Signal Processing"], "engineering",
          "https://www.eurecom.fr/en/teaching/master-networks-and-telecommunication"),
     ]),
    ("IFP School", "Rueil-Malmaison", 0, 6.5,
     "https://www.ifp-school.com/en/programs", [
         ("MSc Oil & Gas Technology", ["Energy Engineering", "Chemical Engineering"], "engineering",
          "https://www.ifp-school.com/en/programs"),
         ("MSc Renewable Energy", ["Energy Engineering", "Sustainability"], "engineering",
          "https://www.ifp-school.com/en/programs"),
     ]),
    ("Bordeaux INP", "Bordeaux", 628, 6.5,
     "https://www.bordeaux-inp.fr/en/bordeaux-inp-programmes", [
         ("MSc Electrical Engineering", ["Electrical Engineering", "Electronics"], "engineering",
          "https://www.bordeaux-inp.fr/en/bordeaux-inp-programmes"),
         ("MSc Computer Science", ["Computer Science"], "cs_ai",
          "https://www.bordeaux-inp.fr/en/bordeaux-inp-programmes"),
     ]),
    ("Grenoble INP", "Grenoble", 628, 6.5,
     "https://www.grenoble-inp.fr/en/academics/masters", [
         ("MSc Industrial Engineering", ["Industrial Engineering"], "engineering",
          "https://www.grenoble-inp.fr/en/academics/masters"),
         ("MSc Computer Science & Networks", ["Computer Science", "Networking"], "cs_ai",
          "https://www.grenoble-inp.fr/en/academics/masters"),
         ("MSc Materials Science & Engineering", ["Materials Science", "Engineering"], "engineering",
          "https://www.grenoble-inp.fr/en/academics/masters"),
     ]),
    ("Toulouse INP", "Toulouse", 628, 6.5,
     "https://www.inp-toulouse.fr/en/study/master-degrees.html", [
         ("MSc Aerospace Engineering", ["Aerospace Engineering"], "engineering",
          "https://www.inp-toulouse.fr/en/study/master-degrees.html"),
         ("MSc Chemical Engineering", ["Chemical Engineering", "Energy"], "engineering",
          "https://www.inp-toulouse.fr/en/study/master-degrees.html"),
     ]),
]

_SCHOLARSHIP_SCHOOLS = {
    "psl", "ip_paris", "paris_saclay", "sorbonne", "grenoble_alpes", "bordeaux",
}

# ── parsers ───────────────────────────────────────────────────────────────────

def parse_france_bachelor_programs(rows: list) -> list[RawProgram]:
    """Parse France bachelors section (cols 12-20 in France sheet)."""
    programs = []
    for row in rows[4:]:
        univ = row[12] if len(row) > 12 else None
        prog_name = row[13] if len(row) > 13 else None
        if _skip_row(univ) or not prog_name:
            continue
        univ = str(univ).strip()
        prog_name = str(prog_name).strip()
        if len(prog_name) < 5:
            continue

        duration = parse_duration(row[14] if len(row) > 14 else None)
        tuition = parse_tuition(row[15] if len(row) > 15 else None)
        city_raw = str(row[16]).split(",")[0].strip() if len(row) > 16 and row[16] else "France"
        req = str(row[17]) if len(row) > 17 and row[17] else ""
        deadline_raw = row[19] if len(row) > 19 else None

        ielts = parse_ielts(req)
        deadline = parse_deadline(deadline_raw)
        category = category_from_name(prog_name)
        fields = [w.strip().title() for w in re.split(r"[,/&—–]", prog_name) if len(w.strip()) > 3][:3]

        programs.append(RawProgram(
            program_name=prog_name,
            university=univ,
            country="France",
            city=city_raw,
            level="bachelor",
            source_name="local_docs",
            source_url="",
            apply_url="",
            category=category,
            field_of_study=fields if fields else [prog_name[:50]],
            duration_years=duration,
            tuition_usd_year=tuition,
            language="English",
            ielts_min=ielts or 6.0,
            gre_required=False,
            gpa_min=None,
            gpa_scale=4.0,
            intake="September",
            deadline=deadline,
            scholarship_available=False,
            description=f"English-taught Bachelor's program at {univ}, France.",
            requirements=[
                "Secondary school diploma / Baccalauréat or equivalent",
                f"English proficiency: {req[:80]}" if req else "English language proficiency",
                "Academic transcripts",
            ],
            qs_ranking=None,
        ))
    return programs


def parse_france_master_programs(rows: list) -> list[RawProgram]:
    """Parse France Masters detailed sheet (specific program + URL rows)."""
    programs = []
    prev_univ = None
    for row in rows[1:]:
        univ_raw = row[0]
        prog_name_raw = row[1] if len(row) > 1 else None
        url_raw = row[2] if len(row) > 2 else None
        tuition_raw = row[3] if len(row) > 3 else None
        req_raw = row[4] if len(row) > 4 else None
        deadline_raw = row[7] if len(row) > 7 else None

        if univ_raw and not _skip_row(univ_raw):
            prev_univ = str(univ_raw).strip()
        univ = prev_univ
        if not univ:
            continue

        prog_name = str(prog_name_raw).strip() if prog_name_raw else None
        if not prog_name or len(prog_name) < 5:
            continue

        url = str(url_raw).strip() if url_raw and str(url_raw).startswith("http") else ""
        tuition = parse_tuition(tuition_raw)
        ielts = parse_ielts(str(req_raw) if req_raw else "")
        deadline = parse_deadline(deadline_raw)
        category = category_from_name(prog_name)
        fields = [w.strip().title() for w in re.split(r"[,/&—–]", prog_name) if len(w.strip()) > 3][:3]

        programs.append(RawProgram(
            program_name=prog_name,
            university=univ,
            country="France",
            city="France",
            level="master",
            source_name="local_docs",
            source_url=url,
            apply_url=url,
            category=category,
            field_of_study=fields if fields else [prog_name[:50]],
            duration_years=2.0,
            tuition_usd_year=tuition,
            language="English",
            ielts_min=ielts or 6.0,
            gre_required=False,
            gpa_min=None,
            gpa_scale=4.0,
            intake="September",
            deadline=deadline,
            scholarship_available=False,
            description=f"English-taught Master's program at {univ}, France.",
            requirements=[
                "Bachelor's degree or equivalent (Licence)",
                f"English requirement: {str(req_raw)[:80]}" if req_raw else "English language proficiency (B2+)",
                "Academic transcripts and motivation letter",
            ],
            qs_ranking=None,
        ))
    return programs


def parse_france_master_universities(rows: list) -> list[RawProgram]:
    """Parse France sheet master section (cols 0-11) — one entry per university."""
    programs = []
    QS_RANKS = {
        "université psl": 28, "paris sciences": 28, "psl": 28,
        "institut polytechnique de paris": 41, "ip paris": 41,
        "paris-saclay": 71, "paris saclay": 71, "paris sacly": 71,
        "sorbonne": 72, "sciences po": 342,
        "paris cité": 300, "paris cite": 300,
        "paris 1": 257, "panthéon-sorbonne": 257,
        "grenoble alpes": 321, "uga": 321,
        "bordeaux": 331, "strasbourg": 351,
        "lyon": 481, "claude bernard": 481,
        "aix-marseille": 241, "montpellier": 401,
        "nantes": 701, "toulouse": 651,
        "lille": 601, "nice": 651,
    }

    for row in rows[4:]:
        univ = row[0]
        if _skip_row(univ):
            continue
        univ = str(univ).strip()
        if not univ or len(univ) < 4:
            continue

        field_hint = str(row[1]).strip() if row[1] else ""
        url = str(row[2]).strip() if len(row) > 2 and row[2] and str(row[2]).startswith("http") else ""
        duration = parse_duration(row[3] if len(row) > 3 else None)
        tuition_raw = row[5] if len(row) > 5 else None
        req = str(row[7]) if len(row) > 7 and row[7] else ""
        deadline_raw = row[10] if len(row) > 10 else None

        tuition = parse_tuition(tuition_raw)
        ielts = parse_ielts(req)
        deadline = parse_deadline(deadline_raw)

        fields = fields_from_abbr(field_hint) if field_hint and field_hint != "None" else ["Various Disciplines"]
        category = category_from_fields(fields) if fields[0] != "Various Disciplines" else "cs_ai"

        # QS rank lookup
        qs = None
        univ_lower = univ.lower()
        for key, rank in QS_RANKS.items():
            if key in univ_lower:
                qs = rank
                break

        scholarship = bool(row[4]) if len(row) > 4 else False

        programs.append(RawProgram(
            program_name=f"English-Taught Master's Programs",
            university=univ,
            country="France",
            city="France",
            level="master",
            source_name="local_docs",
            source_url=url,
            apply_url=url,
            category=category,
            field_of_study=fields[:4],
            duration_years=duration,
            tuition_usd_year=tuition,
            language="English",
            ielts_min=ielts or 6.0,
            gre_required=False,
            gpa_min=None,
            gpa_scale=4.0,
            intake="September",
            deadline=deadline,
            scholarship_available=scholarship,
            description=(
                f"English-taught Master's programs at {univ}, France. "
                f"Key strengths: {field_hint}. See university website for full program catalog."
                if field_hint else
                f"English-taught Master's programs at {univ}, France. "
                "Visit the university website for the full program catalog."
            ),
            requirements=[
                "Bachelor's degree or equivalent (Licence/Baccalauréat+3)",
                f"English proficiency: {req[:80]}" if req else "English proficiency (B2+ / IELTS 6.0+)",
                "Academic transcripts",
                "Statement of purpose / motivation letter",
            ],
            qs_ranking=qs,
        ))
    return programs


def parse_italy_universities(rows: list) -> list[RawProgram]:
    """Parse Italy sheet — one master entry per university."""
    CITY_MAP = {
        "politecnico di milano": "Milan", "università di milano": "Milan",
        "bicocca": "Milan", "bocconi": "Milan", "cattolica": "Milan",
        "insubria": "Como", "bergamo": "Bergamo", "brescia": "Brescia",
        "pavia": "Pavia", "torino": "Turin", "politecnico di torino": "Turin",
        "padova": "Padua", "venezia": "Venice", "ca' foscari": "Venice",
        "verona": "Verona", "bologna": "Bologna", "modena": "Modena",
        "parma": "Parma", "ferrara": "Ferrara", "trieste": "Trieste",
        "udine": "Udine", "genova": "Genoa", "firenze": "Florence",
        "pisa": "Pisa", "siena": "Siena", "perugia": "Perugia",
        "roma": "Rome", "sapienza": "Rome", "tor vergata": "Rome",
        "napoli": "Naples", "salerno": "Salerno", "bari": "Bari",
        "catania": "Catania", "palermo": "Palermo", "messina": "Messina",
        "cagliari": "Cagliari", "sassari": "Sassari",
        "unitrento": "Trento", "trento": "Trento", "bolzano": "Bolzano",
        "aquila": "L'Aquila", "camerino": "Camerino",
    }

    programs = []
    for row in rows[4:]:
        univ = row[0]
        if _skip_row(univ):
            continue
        univ = str(univ).strip()
        if not univ or len(univ) < 4:
            continue

        field_hint = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        url = str(row[2]).strip() if len(row) > 2 and row[2] and str(row[2]).startswith("http") else ""
        duration = parse_duration(row[3] if len(row) > 3 else None)
        scholarship_hint = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        tuition_raw = row[5] if len(row) > 5 else None
        req = str(row[7]) if len(row) > 7 and row[7] else ""
        deadline_raw = row[10] if len(row) > 10 else None

        tuition = parse_tuition(tuition_raw)
        if tuition is None:
            tuition = 300.0  # Italian public university avg enrollment fee (€300)
        ielts = parse_ielts(req)
        deadline = parse_deadline(deadline_raw)

        fields = fields_from_abbr(field_hint) if field_hint else ["Various Disciplines"]
        category = category_from_fields(fields) if fields[0] != "Various Disciplines" else "cs_ai"
        scholarship = bool(scholarship_hint and scholarship_hint.lower() not in ("none", ""))

        city = "Italy"
        univ_lower = univ.lower()
        for key, c in CITY_MAP.items():
            if key in univ_lower:
                city = c
                break

        programs.append(RawProgram(
            program_name="English-Taught Master's Programs",
            university=univ,
            country="Italy",
            city=city,
            level="master",
            source_name="local_docs",
            source_url=url,
            apply_url=url,
            category=category,
            field_of_study=fields[:4],
            duration_years=duration,
            tuition_usd_year=tuition,
            language="English",
            ielts_min=ielts or 6.0,
            gre_required=False,
            gpa_min=None,
            gpa_scale=4.0,
            intake="October",
            deadline=deadline,
            scholarship_available=scholarship,
            description=(
                f"English-taught Master's programs at {univ}, Italy. "
                f"Fields: {field_hint}. "
                f"{'Scholarship available: ' + scholarship_hint + '. ' if scholarship else ''}"
                "Visit the university website for the complete English program catalog."
            ),
            requirements=[
                "Bachelor's degree (Laurea Triennale or equivalent)",
                f"English: {req[:80]}" if req else "English proficiency (B2/IELTS 6.0+)",
                "Credential evaluation may be required for non-EU applicants",
                "Academic transcripts",
            ],
            qs_ranking=None,
        ))
    return programs


def parse_spain_universities(rows: list) -> list[RawProgram]:
    """Parse Spain sheet — master entries (cols 0-11) + bachelor entries (cols 12-20)."""
    programs = []
    CITY_MAP = {
        "madrid": "Madrid", "uc3m": "Madrid", "carlos": "Madrid",
        "autonoma de madrid": "Madrid", "complutense": "Madrid",
        "alcalá": "Alcalá de Henares", "rey juan": "Madrid",
        "barcelona": "Barcelona", "pompeu fabra": "Barcelona",
        "polytectic cat": "Barcelona", "politecnica de cat": "Barcelona",
        "rovira": "Tarragona", "girona": "Girona", "oberta": "Barcelona",
        "valencia": "Valencia", "alicante": "Alicante", "upv": "Valencia",
        "politècnica de valència": "Valencia",
        "granada": "Granada", "sevilla": "Seville", "málaga": "Málaga",
        "pablo de olavide": "Seville",
        "pais vasco": "Bilbao", "navarra": "Pamplona",
        "zaragoza": "Zaragoza", "oviedo": "Oviedo", "cantabria": "Santander",
        "salamanca": "Salamanca", "valladolid": "Valladolid",
        "castilla-la mancha": "Toledo", "murcia": "Murcia",
        "illes balears": "Palma", "la laguna": "Santa Cruz de Tenerife",
        "las palmas": "Las Palmas",
        "santiago de compostela": "Santiago de Compostela",
        "vigo": "Vigo",
    }

    for row in rows[4:]:
        # Masters (cols 0-11)
        univ_m = row[0]
        if not _skip_row(univ_m) and univ_m:
            univ = str(univ_m).strip()
            if len(univ) >= 4:
                url = str(row[2]).strip() if len(row) > 2 and row[2] and str(row[2]).startswith("http") else ""
                duration = parse_duration(row[3] if len(row) > 3 else None)
                tuition = parse_tuition(row[5] if len(row) > 5 else None)
                req = str(row[7]) if len(row) > 7 and row[7] else ""
                deadline_raw = row[10] if len(row) > 10 else None
                ielts = parse_ielts(req)
                deadline = parse_deadline(deadline_raw)

                city = "Spain"
                for key, c in CITY_MAP.items():
                    if key in univ.lower():
                        city = c
                        break

                programs.append(RawProgram(
                    program_name="English-Taught Master's Programs",
                    university=univ,
                    country="Spain",
                    city=city,
                    level="master",
                    source_name="local_docs",
                    source_url=url,
                    apply_url=url,
                    category="cs_ai",
                    field_of_study=["Various Disciplines"],
                    duration_years=duration,
                    tuition_usd_year=tuition,
                    language="English",
                    ielts_min=ielts or 6.0,
                    gre_required=False,
                    gpa_min=None,
                    gpa_scale=4.0,
                    intake="September",
                    deadline=deadline,
                    scholarship_available=bool(row[4]) if len(row) > 4 else False,
                    description=(
                        f"English-taught Master's programs at {univ}, Spain. "
                        "Credential recognition via UNEDasiss required for non-EU applicants. "
                        "Visit the university website for the full English program catalog."
                    ),
                    requirements=[
                        "Bachelor's degree (Grado or equivalent)",
                        f"English: {req[:80]}" if req else "English proficiency (B2/IELTS 6.5+)",
                        "Credential recognition via UNEDasiss (non-EU applicants)",
                        "Academic transcripts",
                    ],
                    qs_ranking=None,
                ))

        # Bachelors (cols 12-20)
        univ_b = row[12] if len(row) > 12 else None
        prog_b = row[13] if len(row) > 13 else None
        if not _skip_row(univ_b) and univ_b and prog_b and str(prog_b).strip():
            univ = str(univ_b).strip()
            prog_name = str(prog_b).strip()
            if len(univ) >= 4 and len(prog_name) >= 5:
                duration = parse_duration(row[14] if len(row) > 14 else None)
                tuition = parse_tuition(row[15] if len(row) > 15 else None)
                city_raw = str(row[16]).split(",")[0].strip() if len(row) > 16 and row[16] else "Spain"
                req = str(row[17]) if len(row) > 17 and row[17] else ""
                deadline_raw = row[19] if len(row) > 19 else None
                ielts = parse_ielts(req)
                deadline = parse_deadline(deadline_raw)
                category = category_from_name(prog_name)
                fields = [w.strip().title() for w in re.split(r"[,/&—–]", prog_name) if len(w.strip()) > 3][:3]

                city = city_raw
                for key, c in CITY_MAP.items():
                    if key in univ.lower():
                        city = c
                        break

                programs.append(RawProgram(
                    program_name=prog_name,
                    university=univ,
                    country="Spain",
                    city=city,
                    level="bachelor",
                    source_name="local_docs",
                    source_url="",
                    apply_url="",
                    category=category,
                    field_of_study=fields if fields else [prog_name[:50]],
                    duration_years=duration,
                    tuition_usd_year=tuition,
                    language="English",
                    ielts_min=ielts or 6.5,
                    gre_required=False,
                    gpa_min=None,
                    gpa_scale=4.0,
                    intake="September",
                    deadline=deadline,
                    scholarship_available=False,
                    description=f"English-taught Bachelor's program at {univ}, Spain.",
                    requirements=[
                        "Secondary school diploma (HSC equiv/12th grade)",
                        f"English: {req[:80]}" if req else "English proficiency (IELTS 6.5+)",
                        "UNEDasiss credential recognition (non-EU applicants)",
                    ],
                    qs_ranking=None,
                ))
    return programs


def parse_france_pdf_programs() -> list[RawProgram]:
    """
    Specific MSc/MA/MBA programs from English_Masters_France_2025_26.pdf.
    Covers 11 public universities (PSL, IP Paris, Paris-Saclay, Sorbonne, …)
    and 8 Grandes Écoles (HEC, INSEAD, ESSEC, ESCP, EDHEC, emlyon, NEOMA, GEM).
    """
    programs = []
    for univ_key, prog_name, fields, category in PDF_PROGRAMS:
        meta = PDF_UNIVERSITIES.get(univ_key)
        if not meta:
            continue
        univ_name, city, qs_rank, tuition_eur, ielts, apply_url = meta
        tuition_usd = float(tuition_eur) if tuition_eur and tuition_eur > 100 else None
        qs_note = f" QS Rank #{qs_rank}." if qs_rank else " Top French institution."
        programs.append(RawProgram(
            program_name=prog_name,
            university=univ_name,
            country="France",
            city=city,
            level="master",
            source_name="local_docs",
            source_url=apply_url,
            apply_url=apply_url,
            category=category,
            field_of_study=fields,
            duration_years=2.0,
            tuition_usd_year=tuition_usd,
            language="English",
            ielts_min=ielts,
            gre_required=False,
            gpa_min=None,
            gpa_scale=4.0,
            intake="September",
            deadline=None,
            scholarship_available=(univ_key in _SCHOLARSHIP_SCHOOLS),
            description=(
                f"English-taught {prog_name} at {univ_name}, France.{qs_note} "
                f"Annual tuition approx. €{tuition_eur:,}."
            ),
            requirements=[
                "Bachelor's degree (Licence or equivalent, 3+ years)",
                f"English proficiency: IELTS {ielts}+ or TOEFL 90+",
                "Academic transcripts and CV",
                "Statement of purpose / motivation letter",
            ],
            qs_ranking=qs_rank,
        ))
    return programs


def parse_france_engineering_schools() -> list[RawProgram]:
    """
    MSc programs at 21 French Grandes Écoles d'Ingénieurs from
    Campus France University List.pdf (CentraleSupélec, INSA network,
    Télécom Paris, ENSAE, ENSTA, EURECOM, Bordeaux/Grenoble/Toulouse INP, etc.).
    """
    programs = []
    for school_name, city, tuition_eur, ielts, listing_url, school_programs in ENGINEERING_SCHOOLS:
        tuition_usd = float(tuition_eur) if tuition_eur > 0 else None
        for prog_tuple in school_programs:
            prog_name, fields, category = prog_tuple[0], prog_tuple[1], prog_tuple[2]
            specific_url = prog_tuple[3] if len(prog_tuple) > 3 else listing_url
            apply_url = specific_url or listing_url
            programs.append(RawProgram(
                program_name=prog_name,
                university=school_name,
                country="France",
                city=city,
                level="master",
                source_name="local_docs",
                source_url=apply_url,
                apply_url=apply_url,
                category=category,
                field_of_study=fields,
                duration_years=2.0,
                tuition_usd_year=tuition_usd,
                language="English",
                ielts_min=ielts,
                gre_required=False,
                gpa_min=None,
                gpa_scale=4.0,
                intake="September",
                deadline=None,
                scholarship_available=True,
                description=(
                    f"English-taught {prog_name} at {school_name}, France. "
                    "Top French engineering school (Grande École d'Ingénieurs). "
                    f"Annual tuition approx. €{tuition_eur:,}."
                ),
                requirements=[
                    "Bachelor's degree in relevant field (3+ years / Licence or equivalent)",
                    f"English proficiency: IELTS {ielts}+ or TOEFL 90+",
                    "Academic transcripts",
                    "Statement of purpose and CV",
                ],
                qs_ranking=None,
            ))
    return programs


# ── web enrichment ────────────────────────────────────────────────────────────

def web_enrich_italian_universities() -> list[RawProgram]:
    """Try to scrape actual program names from key Italian university websites."""
    programs = []
    scrapers = [
        ("Politecnico di Milano", "Milan", "Italy", scrape_polimi_programs),
        ("Università di Bologna", "Bologna", "Italy", scrape_unibo_programs),
        ("Politecnico di Torino", "Turin", "Italy", scrape_polito_programs),
        ("Sapienza Università di Roma", "Rome", "Italy", scrape_sapienza_programs),
    ]

    with httpx.Client(headers=HEADERS, follow_redirects=True) as client:
        for univ, city, country, scraper_fn in scrapers:
            logger.info(f"  Web enriching: {univ}")
            try:
                items = scraper_fn(client)
                seen = set()
                for item in items:
                    name = item["name"].strip()
                    if len(name) < 10 or name in seen:
                        continue
                    seen.add(name)
                    url = item.get("url", "")
                    category = category_from_name(name)
                    fields = [w.strip().title() for w in re.split(r"[,/&—–]", name) if len(w.strip()) > 3][:3]

                    programs.append(RawProgram(
                        program_name=name,
                        university=univ,
                        country=country,
                        city=city,
                        level="master",
                        source_name="local_docs",
                        source_url=url,
                        apply_url=url,
                        category=category,
                        field_of_study=fields if fields else [name[:50]],
                        duration_years=2.0,
                        tuition_usd_year=300.0,
                        language="English",
                        ielts_min=6.0,
                        gre_required=False,
                        gpa_min=None,
                        gpa_scale=4.0,
                        intake="October",
                        deadline=None,
                        scholarship_available=True,
                        description=f"English-taught Master's program at {univ}, {country}.",
                        requirements=[
                            "Bachelor's degree (Laurea Triennale or equivalent)",
                            "English proficiency (B2/IELTS 6.0+)",
                            "Academic transcripts",
                        ],
                        qs_ranking=None,
                    ))
                logger.info(f"    → {len(seen)} programs found via web")
            except Exception as e:
                logger.warning(f"    Web enrich failed for {univ}: {e}")
            time.sleep(1)

    return programs


# ── main crawler ──────────────────────────────────────────────────────────────

class LocalDocsCrawler(BaseProgramCrawler):
    SOURCE_NAME = "local_docs"
    CRAWL_DELAY = 0

    def fetch(self) -> list[RawProgram]:
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed — run: pip install openpyxl")
            return []

        programs: list[RawProgram] = []
        seen: set[str] = set()

        def add(prog_list: list[RawProgram], label: str):
            added = 0
            for p in prog_list:
                fp = make_fp(p.program_name, p.university, p.country, p.level)
                if fp not in seen:
                    seen.add(fp)
                    programs.append(p)
                    added += 1
            logger.info(f"  {label}: {added} programs")

        # ── France PDF programs (hardcoded from English_Masters & Campus France PDFs) ──
        add(parse_france_pdf_programs(), "France — PDF masters (19 universities/schools)")
        add(parse_france_engineering_schools(), "France — engineering Grandes Écoles (21 schools)")

        # ── France ──────────────────────────────────────────────────────────
        france_path = DOCS_DIR / "France University Database.xlsx"
        if france_path.exists():
            wb = openpyxl.load_workbook(france_path, read_only=True)

            if "France" in wb.sheetnames:
                rows = list(wb["France"].iter_rows(values_only=True))
                add(parse_france_bachelor_programs(rows), "France bachelors (Excel)")
                add(parse_france_master_universities(rows), "France masters — university level")

            if "France Masters" in wb.sheetnames:
                rows_m = list(wb["France Masters"].iter_rows(values_only=True))
                add(parse_france_master_programs(rows_m), "France masters — specific programs")
        else:
            logger.warning(f"France Excel not found: {france_path}")

        # ── Italy ────────────────────────────────────────────────────────────
        italy_path = DOCS_DIR / "Italy University Database.xlsx"
        if italy_path.exists():
            wb_it = openpyxl.load_workbook(italy_path, read_only=True)
            if "Italy" in wb_it.sheetnames:
                rows_it = list(wb_it["Italy"].iter_rows(values_only=True))
                add(parse_italy_universities(rows_it), "Italy masters — university level")
        else:
            logger.warning(f"Italy Excel not found: {italy_path}")

        # ── Spain ────────────────────────────────────────────────────────────
        spain_path = DOCS_DIR / "Spain University Database.xlsx"
        if spain_path.exists():
            wb_es = openpyxl.load_workbook(spain_path, read_only=True)
            if "Spain" in wb_es.sheetnames:
                rows_es = list(wb_es["Spain"].iter_rows(values_only=True))
                add(parse_spain_universities(rows_es), "Spain masters + bachelors")
        else:
            logger.warning(f"Spain Excel not found: {spain_path}")

        # ── Web enrichment for key Italian universities ──────────────────────
        logger.info("  Web enriching Italian university program pages...")
        web_programs = web_enrich_italian_universities()
        add(web_programs, "Italy — web-enriched programs")

        logger.info(f"LocalDocs total: {len(programs)} programs")
        return programs


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(name)s] %(levelname)s — %(message)s")
    crawler = LocalDocsCrawler()
    items = crawler.fetch()
    print(f"\nFetched {len(items)} programs total")
    from collections import Counter
    for (country, level), count in Counter((p.country, p.level) for p in items).most_common():
        print(f"  {country} / {level}: {count}")
    if "--save" in sys.argv:
        new, updated = crawler.run()
        print(f"Saved: {new} new, {updated} updated")
