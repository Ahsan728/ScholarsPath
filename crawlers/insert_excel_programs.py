#!/usr/bin/env python3
"""
Re-process the priority-country Excel files (France / Italy / Spain) to
extract the SECOND data section we missed on the first pass.

Each row in these files has two halves:
- cols 0-11: university-level metadata (name, website, scholarship URL,
  general application info). Already ingested by an earlier crawler.
- cols 12-20: specific PROGRAM details (university name, program name,
  duration, tuition fees, location, requirements, deadline, status).
  This is what we extract here.

Output goes to masters_programs, gated by Phase 0 validators.

Run:
  python crawlers/insert_excel_programs.py --dry-run
  python crawlers/insert_excel_programs.py
"""

import argparse
import hashlib
import os
import re
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import httpx
import openpyxl
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env.local"))

from crawler_logger import CrawlerRun
from aggregator_hosts import is_aggregator_host
from domain_classifier import classify as classify_domain

SB_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SB_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
SB_H   = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}",
          "Content-Type": "application/json", "Prefer": "return=minimal"}
SB_R   = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}

FILES = [
    ("Documents/France University Database.xlsx", "France"),
    ("Documents/Italy University Database.xlsx",  "Italy"),
    ("Documents/Spain University Database.xlsx",  "Spain"),
]


def parse_ielts(req_text: str) -> float | None:
    """Find the minimum IELTS band from a requirements string."""
    if not req_text:
        return None
    m = re.search(r"IELTS\s*(\d(?:\.\d)?)", req_text, re.I)
    if m:
        try: return float(m.group(1))
        except ValueError: pass
    return None


def parse_tuition(t: str) -> float | None:
    """Best-effort EUR tuition extraction. Returns the lower bound as float."""
    if not t: return None
    s = t.replace(",", "").replace("\xa0", " ")
    m = re.search(r"€\s*(\d{2,6})", s)
    if m:
        try: return float(m.group(1))
        except ValueError: pass
    m = re.search(r"(\d{2,6})\s*(?:EUR|€)", s, re.I)
    if m:
        try: return float(m.group(1))
        except ValueError: pass
    return None


def parse_duration(d: str) -> float | None:
    if not d: return None
    s = str(d).lower()
    if "1 year" in s or "1-year" in s or "1yr" in s: return 1.0
    if "2 year" in s or "2-year" in s or "2yr" in s: return 2.0
    if "3 year" in s or "3-year" in s or "3yr" in s: return 3.0
    if "4 year" in s or "4-year" in s or "4yr" in s: return 4.0
    if "5 year" in s or "5-year" in s or "5yr" in s: return 5.0
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:years?|yr)", s, re.I)
    if m:
        try: return float(m.group(1))
        except ValueError: pass
    return None


def detect_level(program_name: str, duration_years: float | None) -> str:
    if not program_name:
        return "master"
    pn = program_name.lower()
    if any(k in pn for k in ["bba", "bachelor", "ba in", "bs in", "b.sc", "bsc",
                              "undergraduate", "licence"]):
        return "bachelor"
    if "phd" in pn or "doctoral" in pn:
        return "master"  # masters_programs only allows bachelor/master
    if duration_years and duration_years >= 3.5:
        return "bachelor"
    return "master"


def load_existing_fingerprints(country: str) -> set[str]:
    out = set()
    offset = 0
    while True:
        r = httpx.get(
            f"{SB_URL}/rest/v1/masters_programs",
            headers=SB_R,
            params={"select": "fingerprint", "country": f"eq.{country}",
                    "limit": "1000", "offset": str(offset)},
            timeout=20,
        )
        rows = r.json() if r.status_code == 200 else []
        for row in rows:
            f = row.get("fingerprint")
            if f: out.add(f)
        if len(rows) < 1000: break
        offset += 1000
    return out


def load_university_url_map(country: str) -> dict[str, str]:
    """Build a fuzzy university-name → apply_url root map from existing
    programs we've already validated. Used to assign a proper URL to
    new rows that come without one."""
    out: dict[str, str] = {}
    offset = 0
    while True:
        r = httpx.get(
            f"{SB_URL}/rest/v1/masters_programs",
            headers=SB_R,
            params={"select": "university,apply_url",
                    "country": f"eq.{country}",
                    "is_active": "eq.true",
                    "url_status": "eq.ok",
                    "limit": "1000", "offset": str(offset)},
            timeout=20,
        )
        rows = r.json() if r.status_code == 200 else []
        for row in rows:
            uni = (row.get("university") or "").lower().strip()
            url = (row.get("apply_url") or "").strip()
            if not uni or not url.startswith("http"):
                continue
            # Strip path to get domain root
            from urllib.parse import urlparse
            try:
                p = urlparse(url)
                root = f"{p.scheme}://{p.netloc}"
                # Normalize uni name — strip parenthesized acronyms etc.
                key = re.sub(r"\(.*?\)", "", uni)
                key = re.sub(r"\s+", " ", key).strip()
                if key and key not in out:
                    out[key] = root
            except Exception:
                continue
        if len(rows) < 1000: break
        offset += 1000
    return out


def find_url_for(uni_name: str, url_map: dict[str, str]) -> str | None:
    """Return the best-matching university URL root (or None)."""
    if not uni_name:
        return None
    key = uni_name.lower()
    key = re.sub(r"\(.*?\)", "", key)  # strip "(UC3M)" etc.
    key = re.sub(r"\s+", " ", key).strip()
    if key in url_map:
        return url_map[key]
    # Substring match — uni acronym in stored name or vice versa
    for stored_key, url in url_map.items():
        if not stored_key or not key:
            continue
        if key in stored_key or stored_key in key:
            return url
        # Token overlap
        key_toks = set(t for t in key.split() if len(t) > 3)
        stored_toks = set(t for t in stored_key.split() if len(t) > 3)
        if key_toks and stored_toks and len(key_toks & stored_toks) >= 2:
            return url
    return None


def fingerprint(name: str, university: str, country: str, level: str) -> str:
    raw = f"{name.lower().strip()}|{university.lower().strip()}|{country.lower()}|{level}"
    return hashlib.sha256(raw.encode()).hexdigest()


def find_program_columns(headers: list, country: str) -> dict:
    """Discover the column indices for the second-section program data.
    We look for a SECOND occurrence of 'University' followed by
    'Program Name'."""
    idx = {}
    seen_uni = False
    for i, h in enumerate(headers):
        if not h: continue
        h_low = str(h).strip().lower()
        if h_low == "university":
            if not seen_uni:
                seen_uni = True
            else:
                idx["university"] = i
        if "university" in idx:
            if "program" in h_low and "name" in h_low and "program_name" not in idx:
                idx["program_name"] = i
            elif "duration" in h_low and "duration" not in idx:
                idx["duration"] = i
            elif "tuition" in h_low and "tuition" not in idx:
                idx["tuition"] = i
            elif "location" in h_low and "location" not in idx:
                idx["location"] = i
            elif "requirement" in h_low and "requirements" not in idx:
                idx["requirements"] = i
            elif ("application" in h_low or "app." in h_low) and "fee" in h_low and "app_fee" not in idx:
                idx["app_fee"] = i
            elif "deadline" in h_low and "deadline" not in idx:
                idx["deadline"] = i
            elif "status" in h_low and "status" not in idx:
                idx["status"] = i
    return idx


def insert(record: dict, existing: set[str], dry_run: bool) -> bool:
    fp = record["fingerprint"]
    if fp in existing:
        return False
    if dry_run:
        print(f"  WOULD INSERT: [{record['country']:7s}] {record['university'][:28]:28s} | {record['program_name'][:50]}",
              flush=True)
        return True
    r = httpx.post(f"{SB_URL}/rest/v1/masters_programs",
                   headers=SB_H, json=record, timeout=15)
    if r.status_code in (200, 201, 204):
        existing.add(fp)
        return True
    print(f"  FAIL {r.status_code}: {r.text[:200]}", flush=True)
    return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    with CrawlerRun("excel_programs_reingest",
                    params={"dry_run": args.dry_run}) as run:
        grand_total = 0
        for path, country in FILES:
            full = os.path.join(os.path.dirname(__file__), "..", path)
            if not os.path.exists(full):
                print(f"SKIP missing: {full}", flush=True)
                continue
            print(f"\n=== {path} ({country}) ===", flush=True)
            wb = openpyxl.load_workbook(full, read_only=True)
            ws = wb.active

            headers = None
            for row in ws.iter_rows(values_only=True, max_row=10):
                if row[0] == "University":
                    headers = list(row)
                    break
            if not headers:
                print("  no header row found", flush=True)
                wb.close()
                continue

            idx = find_program_columns(headers, country)
            if not idx.get("university") or not idx.get("program_name"):
                print(f"  no program section found in {path}", flush=True)
                wb.close()
                continue
            print(f"  program section indices: {idx}", flush=True)

            existing = load_existing_fingerprints(country)
            url_map = load_university_url_map(country)
            print(f"  existing {country} fingerprints: {len(existing)} | "
                  f"uni→URL map: {len(url_map)}", flush=True)

            inserted = 0
            for row in ws.iter_rows(values_only=True):
                if not row or row[0] == "University":
                    continue
                uni = str(row[idx["university"]] or "").strip() if idx["university"] < len(row) else ""
                pname = str(row[idx["program_name"]] or "").strip() if idx["program_name"] < len(row) else ""
                if not uni or not pname:
                    continue
                # Header rows can leak through; skip if uni equals "University"
                if uni.lower() == "university" or pname.lower() == "program name":
                    continue

                duration_text = str(row[idx["duration"]] or "").strip() if idx.get("duration") and idx["duration"] < len(row) else ""
                tuition_text  = str(row[idx["tuition"]]  or "").strip() if idx.get("tuition")  and idx["tuition"] < len(row) else ""
                location      = str(row[idx["location"]] or "").strip() if idx.get("location") and idx["location"] < len(row) else ""
                requirements  = str(row[idx["requirements"]] or "").strip() if idx.get("requirements") and idx["requirements"] < len(row) else ""
                deadline      = str(row[idx["deadline"]] or "").strip() if idx.get("deadline") and idx["deadline"] < len(row) else ""

                duration_years = parse_duration(duration_text)
                ielts_min = parse_ielts(requirements)
                tuition_usd = parse_tuition(tuition_text)
                level = detect_level(pname, duration_years)

                fp = fingerprint(pname, uni, country, level)
                if fp in existing:
                    continue

                category = classify_domain([pname], pname)
                city = (location.split(",")[0].strip() if location else country)[:100]

                # Try to assign a real URL for this university
                apply_url = find_url_for(uni, url_map)
                if not apply_url:
                    # Skip rows where we can't find a real URL — Phase 0
                    # would reject them anyway.
                    continue

                record = {
                    "program_name":    pname[:300],
                    "university":      uni[:300],
                    "country":         country,
                    "city":            city,
                    "level":           level,
                    "duration_years":  duration_years or (2.0 if level == "master" else 3.0),
                    "tuition_usd_year": tuition_usd,
                    "language":        "English",
                    "field_of_study":  [],
                    "category":        category,
                    "ielts_min":       ielts_min,
                    "gre_required":    False,
                    "gpa_min":         None,
                    "gpa_scale":       4.0,
                    "intake":          "Fall/Spring",
                    "deadline":        None,
                    "scholarship_available": False,
                    "description":     f"{pname} at {uni}, {country}. Listed in the curated {country} university database.",
                    "requirements":    [requirements] if requirements else [],
                    "apply_url":       apply_url,
                    "source_url":      path,
                    "source_name":     f"excel_{country.lower()}_v2",
                    "is_active":       True,
                    "fingerprint":     fp,
                }
                if insert(record, existing, args.dry_run):
                    inserted += 1
                    run.ok()
                else:
                    run.skipped()
            print(f"  {country}: +{inserted} programs", flush=True)
            grand_total += inserted
            wb.close()

        run.summary = {"total_inserted": grand_total}
        print(f"\nDONE: {grand_total} programs inserted", flush=True)


if __name__ == "__main__":
    main()
