#!/usr/bin/env python3
"""
Ingest the France Research Institutes Directory from
Documents/france_research_institutes.xlsx.

Two productive sheets:
- 'National Organizations' (13 institutes): CNRS, Inserm, INRAE, CEA,
  Inria, IRD, IFREMER, ONERA, IFP, BRGM, IGN, CIRAD, Institut Pasteur,
  Institut Curie.
- 'Named Institutes' (52 institutes): flagship labs grouped by domain
  with Acronym, Full Name, City, Affiliation, Website.

For each institute with a website, we insert:
1. one 'phd' opportunity row in discovered_opportunities representing
   their open PhD/postdoc positions (rolling deadline);
2. one row in opportunity_sources so the Discoverer crawls their
   careers/jobs page on its monthly sweep.

Pattern mirrors crawlers/ingest_spain_research_institutes.py.

Run:
  python crawlers/ingest_france_research_institutes.py --dry-run
  python crawlers/ingest_france_research_institutes.py
"""

import argparse
import hashlib
import os
import sys
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import httpx
import openpyxl
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env.local"))

from crawler_logger import CrawlerRun
from aggregator_hosts import is_aggregator_host
from domain_classifier import classify as classify_domain

SB_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SB_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
SB_H   = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}",
          "Content-Type": "application/json", "Prefer": "return=minimal"}
SB_R   = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}

XLSX = os.path.join(os.path.dirname(__file__), "..", "Documents",
                    "france_research_institutes.xlsx")


def load_institutes() -> list[dict]:
    """Parse both useful sheets into a uniform list of dicts."""
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    out: list[dict] = []

    # Sheet: National Organizations
    ws = wb["National Organizations"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[2:]:  # skip title + header
        if not r or not r[0]: continue
        acr, fullname, scope, website = (r + (None,) * 4)[:4]
        if not website: continue
        out.append({
            "acronym":  str(acr or "").strip(),
            "name":     str(fullname or "").strip(),
            "domain":   str(scope or "").strip(),
            "city":     "",
            "affiliation": "",
            "website":  str(website).strip(),
            "tier":     "national_organization",
        })

    # Sheet: Named Institutes
    ws = wb["Named Institutes"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[2:]:  # skip title + header
        if not r or not r[0]: continue
        dom, acr, fullname, city, aff, website, _ = (r + (None,) * 7)[:7]
        if not website: continue
        out.append({
            "acronym":  str(acr or "").strip(),
            "name":     str(fullname or "").strip(),
            "domain":   str(dom or "").strip(),
            "city":     str(city or "").strip(),
            "affiliation": str(aff or "").strip(),
            "website":  str(website).strip(),
            "tier":     "named_institute",
        })

    wb.close()
    # dedup by acronym + name + website (sheets shouldn't overlap but be safe)
    seen = set()
    uniq = []
    for r in out:
        key = (r["acronym"].lower(), r["name"].lower(), r["website"].lower())
        if key in seen: continue
        seen.add(key)
        uniq.append(r)
    return uniq


def url_with_scheme(domain: str) -> str | None:
    domain = (domain or "").strip().lower()
    if not domain or domain in ("none", "n/a", "-"):
        return None
    if domain.startswith("http://") or domain.startswith("https://"):
        return domain
    return f"https://{domain}"


def insert_opportunity(inst: dict, dry_run: bool) -> bool:
    url = url_with_scheme(inst["website"])
    if not url or is_aggregator_host(url):
        return False

    title = f"PhD / Postdoc positions at {inst['acronym']} — {inst['name']}"
    parts = [f"{inst['acronym']} ({inst['name']}) is a French research "
             f"{'national organization' if inst['tier']=='national_organization' else 'institute / lab'}"]
    if inst["domain"]:
        parts.append(f"focused on {inst['domain']}")
    if inst["city"]:
        parts.append(f"based in {inst['city']}")
    if inst["affiliation"]:
        parts.append(f"({inst['affiliation']})")
    desc = ", ".join(parts).rstrip(",") + (
        ". Hires PhD researchers and postdoctoral fellows on open calls."
    )

    content_hash = hashlib.sha256(
        f"france-ri|{inst['acronym']}|{inst['name']}".encode()
    ).hexdigest()
    category = classify_domain([inst["domain"]], title)

    record = {
        "source_url":     "Documents/france_research_institutes.xlsx",
        "prompt_version": "france-research-institutes-v1",
        "content_hash":   content_hash,
        "type":           "phd",
        "title":          title[:300],
        "description":    desc[:2000],
        "university":     inst["name"][:300] or inst["acronym"][:300],
        "country":        "France",
        "degree_level":   "phd",
        "field_of_study": [inst["domain"]] if inst["domain"] else [],
        "category":       category,
        "amount_text":    None,
        "funding_type":   None,
        "eligibility_text": None,
        "eligible_nations": ["ALL"],
        "ineligible_nations": [],
        "deadline":       None,
        "deadline_text":  "Rolling / per-call",
        "intake":         None,
        "apply_url":      url,
        "is_active":      True,
        "last_seen_at":   datetime.now(timezone.utc).isoformat(),
    }
    if dry_run:
        print(f"  WOULD INSERT: {title[:80]}", flush=True)
        return True

    ex = httpx.get(
        f"{SB_URL}/rest/v1/discovered_opportunities",
        headers=SB_R,
        params={"select": "id", "content_hash": f"eq.{content_hash}", "limit": "1"},
        timeout=10,
    )
    if ex.status_code == 200 and ex.json():
        return False
    r = httpx.post(f"{SB_URL}/rest/v1/discovered_opportunities",
                   headers=SB_H, json=record, timeout=15)
    return r.status_code in (200, 201, 204)


def insert_source(inst: dict, dry_run: bool) -> bool:
    url = url_with_scheme(inst["website"])
    if not url or is_aggregator_host(url):
        return False
    if dry_run:
        return True

    ex = httpx.get(
        f"{SB_URL}/rest/v1/opportunity_sources",
        headers=SB_R,
        params={"select": "id", "url": f"eq.{url}", "limit": "1"},
        timeout=10,
    )
    if ex.status_code == 200 and ex.json():
        return False

    scope = ("funding_body" if inst["tier"] == "national_organization"
             else "research_institute")
    notes_extra = f"Affiliation: {inst['affiliation']}. " if inst["affiliation"] else ""
    record = {
        "url":     url,
        "country": "France",
        "scope":   scope,
        "title":   f"{inst['acronym']} — {inst['name']} (France research institute)"[:300],
        "added_by": "france_research_institutes_v1",
        "notes":   f"Domain: {inst['domain']}. City: {inst['city']}. "
                   f"{notes_extra}Source: "
                   f"Documents/france_research_institutes.xlsx",
    }
    r = httpx.post(f"{SB_URL}/rest/v1/opportunity_sources",
                   headers=SB_H, json=record, timeout=15)
    return r.status_code in (200, 201, 204)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--skip-source", action="store_true")
    ap.add_argument("--skip-opp", action="store_true")
    args = ap.parse_args()

    institutes = load_institutes()
    print(f"Loaded {len(institutes)} institutes from spreadsheet", flush=True)

    with CrawlerRun("france_research_institutes_ingest",
                    params={"dry_run": args.dry_run,
                            "skip_source": args.skip_source,
                            "skip_opp": args.skip_opp}) as run:
        run.set_total(len(institutes))
        opps_written = srcs_written = skipped = 0
        for inst in institutes:
            if not inst["website"]:
                skipped += 1
                run.skipped()
                continue
            opp_ok = src_ok = False
            if not args.skip_opp:
                opp_ok = insert_opportunity(inst, args.dry_run)
                if opp_ok: opps_written += 1
            if not args.skip_source:
                src_ok = insert_source(inst, args.dry_run)
                if src_ok: srcs_written += 1
            if opp_ok or src_ok:
                run.ok()
                if not args.dry_run:
                    print(f"  + {inst['acronym']:14s} {inst['name'][:55]}",
                          flush=True)
            else:
                run.skipped()
                skipped += 1

        run.summary = {
            "institutes_total": len(institutes),
            "opportunities_written": opps_written,
            "sources_written": srcs_written,
            "skipped": skipped,
        }
        print(f"\nDONE: opportunities={opps_written}, sources={srcs_written}, "
              f"skipped={skipped} (of {len(institutes)} institutes)",
              flush=True)


if __name__ == "__main__":
    main()
