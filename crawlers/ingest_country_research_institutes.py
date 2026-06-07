#!/usr/bin/env python3
"""
Generic research-institutes ingester for the country Excel files in
Documents/<country>_research_institutes.xlsx.

Six countries, all share the same two productive sheets:

  Sheet A — national orgs/funders/bodies
    cols: Name | Type | Scope/role | Website                  (4 cols)
  Sheet B — named institutes / labs
    cols: Domain | Institute | City | Network/Host | Website | Notes  (6 cols)

The sheet *names* differ per country, so CONFIG maps each country slug
to its actual sheet names. Insertion logic is identical to
crawlers/ingest_france_research_institutes.py and
crawlers/ingest_spain_research_institutes.py: one phd opportunity per
institute (rolling deadline) + one row in opportunity_sources for the
Discoverer to crawl.

Run:
  python crawlers/ingest_country_research_institutes.py --country hungary --dry-run
  python crawlers/ingest_country_research_institutes.py --country all
"""

import argparse
import hashlib
import os
import sys
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import httpx
import openpyxl
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env.local"))

from crawler_logger import CrawlerRun
from aggregator_hosts import is_aggregator_host
from domain_classifier import classify as classify_domain

SB_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SB_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
SB_H   = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}",
          "Content-Type": "application/json", "Prefer": "return=minimal"}
SB_R   = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}

# slug -> (display country, sheet_a_name, sheet_b_name, sheet_a_kind)
# sheet_a_kind is 'national_organization' or 'funder' — used to set
# opportunity_sources.scope.
CONFIG = {
    "hungary": ("Hungary", "Universities & Funders", "Named Institutes", "funder"),
    "germany": ("Germany", "National Organizations",  "Named Institutes", "national_organization"),
    "netherlands": ("Netherlands", "Universities & Funders", "Named Institutes", "funder"),
    "sweden": ("Sweden", "Universities & Funders", "Institutes & Infrastructure", "funder"),
    "belgium": ("Belgium", "Bodies, Funders & SRCs", "Named Institutes", "funder"),
    "italy": ("Italy", "National Bodies & Schools", "Named Institutes", "national_organization"),
}


def load_institutes(slug: str) -> list[dict]:
    """Parse the two productive sheets for one country."""
    country, sheet_a, sheet_b, sheet_a_kind = CONFIG[slug]
    xlsx = os.path.join(os.path.dirname(__file__), "..",
                        "Documents", f"{slug}_research_institutes.xlsx")
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    out: list[dict] = []

    # Sheet A — orgs/funders (4 cols: Name, Type, Scope, Website)
    if sheet_a in wb.sheetnames:
        ws = wb[sheet_a]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[2:]:  # skip title + header
            if not r or not r[0]: continue
            name, typ, scope, website = (r + (None,) * 4)[:4]
            if not website: continue
            out.append({
                "acronym":     "",
                "name":        str(name or "").strip(),
                "domain":      str(scope or "").strip(),
                "city":        "",
                "affiliation": str(typ or "").strip(),
                "website":     str(website).strip(),
                "tier":        sheet_a_kind,
            })

    # Sheet B — named institutes (6 cols: Domain, Institute, City, Org, Website, Notes)
    if sheet_b in wb.sheetnames:
        ws = wb[sheet_b]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[2:]:
            if not r or not r[0]: continue
            dom, inst, city, host, website, _ = (r + (None,) * 6)[:6]
            if not website: continue
            out.append({
                "acronym":     "",
                "name":        str(inst or "").strip(),
                "domain":      str(dom or "").strip(),
                "city":        str(city or "").strip(),
                "affiliation": str(host or "").strip(),
                "website":     str(website).strip(),
                "tier":        "named_institute",
            })
    wb.close()

    # dedup (name + website) — be safe if a row appears in both sheets
    seen, uniq = set(), []
    for r in out:
        key = (r["name"].lower(), r["website"].lower())
        if key in seen: continue
        seen.add(key)
        uniq.append(r)
    return uniq


def url_with_scheme(domain: str) -> str | None:
    domain = (domain or "").strip().lower()
    if not domain or domain in ("none", "n/a", "-"):
        return None
    if domain.startswith("http://") or domain.startswith("https://"):
        return domain
    return f"https://{domain}"


def insert_opportunity(inst: dict, country: str, slug: str,
                       dry_run: bool) -> bool:
    url = url_with_scheme(inst["website"])
    if not url or is_aggregator_host(url):
        return False
    label = inst["name"] or inst.get("acronym") or ""
    title = f"PhD / Postdoc positions at {label}"

    parts = [f"{label} is a {country} research "
             f"{'organization' if inst['tier']!='named_institute' else 'institute / lab'}"]
    if inst["domain"]: parts.append(f"focused on {inst['domain']}")
    if inst["city"]:   parts.append(f"based in {inst['city']}")
    if inst["affiliation"]: parts.append(f"({inst['affiliation']})")
    desc = ", ".join(parts).rstrip(",") + (
        ". Hires PhD researchers and postdoctoral fellows on open calls."
    )

    content_hash = hashlib.sha256(
        f"{slug}-ri|{label}|{url}".encode()
    ).hexdigest()
    category = classify_domain([inst["domain"]], title)

    record = {
        "source_url":     f"Documents/{slug}_research_institutes.xlsx",
        "prompt_version": f"{slug}-research-institutes-v1",
        "content_hash":   content_hash,
        "type":           "phd",
        "title":          title[:300],
        "description":    desc[:2000],
        "university":     label[:300],
        "country":        country,
        "degree_level":   "phd",
        "field_of_study": [inst["domain"]] if inst["domain"] else [],
        "category":       category,
        "amount_text":    None,
        "funding_type":   None,
        "eligibility_text": None,
        "eligible_nations": ["ALL"],
        "ineligible_nations": [],
        "deadline":       None,
        "deadline_text":  "Rolling / per-call",
        "intake":         None,
        "apply_url":      url,
        "is_active":      True,
        "last_seen_at":   datetime.now(timezone.utc).isoformat(),
    }
    if dry_run:
        print(f"  WOULD INSERT: {title[:80]}", flush=True)
        return True

    ex = httpx.get(
        f"{SB_URL}/rest/v1/discovered_opportunities",
        headers=SB_R,
        params={"select": "id", "content_hash": f"eq.{content_hash}", "limit": "1"},
        timeout=10,
    )
    if ex.status_code == 200 and ex.json():
        return False
    r = httpx.post(f"{SB_URL}/rest/v1/discovered_opportunities",
                   headers=SB_H, json=record, timeout=15)
    return r.status_code in (200, 201, 204)


def insert_source(inst: dict, country: str, slug: str, dry_run: bool) -> bool:
    url = url_with_scheme(inst["website"])
    if not url or is_aggregator_host(url):
        return False
    if dry_run:
        return True

    ex = httpx.get(
        f"{SB_URL}/rest/v1/opportunity_sources",
        headers=SB_R,
        params={"select": "id", "url": f"eq.{url}", "limit": "1"},
        timeout=10,
    )
    if ex.status_code == 200 and ex.json():
        return False

    scope = "research_institute" if inst["tier"] == "named_institute" else (
        "funding_body" if inst["tier"] != "national_organization" else "funding_body"
    )
    label = inst["name"] or inst.get("acronym") or ""
    record = {
        "url":     url,
        "country": country,
        "scope":   scope,
        "title":   f"{label} ({country} research institute)"[:300],
        "added_by": f"{slug}_research_institutes_v1",
        "notes":   f"Domain: {inst['domain']}. City: {inst['city']}. "
                   f"Affiliation: {inst['affiliation']}. Source: "
                   f"Documents/{slug}_research_institutes.xlsx",
    }
    r = httpx.post(f"{SB_URL}/rest/v1/opportunity_sources",
                   headers=SB_H, json=record, timeout=15)
    return r.status_code in (200, 201, 204)


def run_one(slug: str, dry_run: bool, skip_source: bool, skip_opp: bool):
    country, _, _, _ = CONFIG[slug]
    institutes = load_institutes(slug)
    print(f"\n=== {country} ({slug}) — {len(institutes)} institutes ===",
          flush=True)
    with CrawlerRun(f"{slug}_research_institutes_ingest",
                    params={"dry_run": dry_run,
                            "skip_source": skip_source,
                            "skip_opp": skip_opp}) as run:
        run.set_total(len(institutes))
        opps = srcs = skipped = 0
        for inst in institutes:
            if not inst["website"]:
                skipped += 1; run.skipped(); continue
            opp_ok = src_ok = False
            if not skip_opp:
                opp_ok = insert_opportunity(inst, country, slug, dry_run)
                if opp_ok: opps += 1
            if not skip_source:
                src_ok = insert_source(inst, country, slug, dry_run)
                if src_ok: srcs += 1
            if opp_ok or src_ok:
                run.ok()
                if not dry_run:
                    label = inst["name"] or inst.get("acronym") or ""
                    print(f"  + {label[:55]}", flush=True)
            else:
                run.skipped(); skipped += 1
        run.summary = {
            "country": country,
            "institutes_total": len(institutes),
            "opportunities_written": opps,
            "sources_written": srcs,
            "skipped": skipped,
        }
        print(f"  {country}: opps={opps}, sources={srcs}, skipped={skipped}",
              flush=True)
        return opps, srcs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--country", required=True,
                    help=f"One of {list(CONFIG.keys())} or 'all'")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--skip-source", action="store_true")
    ap.add_argument("--skip-opp", action="store_true")
    args = ap.parse_args()

    if args.country == "all":
        slugs = list(CONFIG.keys())
    elif args.country in CONFIG:
        slugs = [args.country]
    else:
        print(f"Unknown country '{args.country}'. Use {list(CONFIG.keys())} or 'all'")
        sys.exit(1)

    total_opps = total_srcs = 0
    for slug in slugs:
        o, s = run_one(slug, args.dry_run, args.skip_source, args.skip_opp)
        total_opps += o
        total_srcs += s

    print(f"\nGRAND TOTAL: opportunities={total_opps}, sources={total_srcs}",
          flush=True)


if __name__ == "__main__":
    main()
